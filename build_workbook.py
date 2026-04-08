from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import re
from typing import List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "packaging_skills_matrix.xlsx"
LINE_SOURCE_FILE = "PackagingLines.xlsx"
EMPLOYEE_SOURCE_FILE = "EMPLOYEE ROSTER.xlsx"

SKILLS = [
    ("Safety and Compliance", "C01", "Applies LOTO correctly for the task being performed", "Performs LOTO without missing steps and verifies zero energy before work begins", "Yes", 3, ""),
    ("Safety and Compliance", "C02", "Identifies emergency stops and safety devices on the assigned line", "Can locate and explain the purpose of major safety devices without prompting", "Yes", 3, ""),
    ("Safety and Compliance", "C03", "Uses proper PPE and follows GMP and GDP during active work", "Uses correct PPE and maintains compliant practices during normal work without reminders", "Yes", 3, ""),
    ("Line and Equipment Understanding", "C04", "Explains line flow from start to finish", "Can walk the line and explain the purpose of each major section in the process", "No", 3, ""),
    ("Line and Equipment Understanding", "C05", "Identifies major components and their function", "Correctly identifies major machine components and explains their basic function", "No", 3, ""),
    ("Line and Equipment Understanding", "C06", "Navigates HMI screens needed for operation", "Uses the HMI to monitor status and perform normal operating actions without help", "No", 3, ""),
    ("Startup and Pre Run Checks", "C07", "Performs pre run checks completely", "Completes all required pre run checks without skipping steps", "Yes", 3, ""),
    ("Startup and Pre Run Checks", "C08", "Verifies correct components against the work order", "Confirms components match the work order and catches mismatches before startup", "Yes", 3, ""),
    ("Startup and Pre Run Checks", "C09", "Confirms line readiness before startup", "Verifies line setup, materials, and readiness before startup without prompting", "Yes", 3, ""),
    ("Running the Line", "C10", "Maintains steady line operation", "Runs the line with normal stability and avoids unnecessary stops caused by operator error", "No", 3, ""),
    ("Running the Line", "C11", "Makes basic adjustments to maintain flow", "Performs normal operating adjustments to keep the line flowing without needing frequent help", "No", 3, ""),
    ("Running the Line", "C12", "Communicates issues before they escalate", "Recognizes developing issues and communicates early with the right level of detail", "No", 3, ""),
    ("In Process Quality", "C13", "Verifies coding against the work order", "Confirms coding matches the work order before release and recognizes obvious coding errors", "Yes", 3, ""),
    ("In Process Quality", "C14", "Performs FAI verification correctly", "Completes required FAI verification steps accurately without missing key checks", "Yes", 3, ""),
    ("In Process Quality", "C15", "Identifies common defects and responds appropriately", "Recognizes common product or package defects and takes the correct next step", "Yes", 3, ""),
    ("Troubleshooting and Mechanical Skill", "C16", "Clears jams safely", "Clears common jams safely without causing additional issues", "No", 3, ""),
    ("Troubleshooting and Mechanical Skill", "C17", "Performs minor mechanical adjustments", "Completes routine minor adjustments correctly and knows when the issue is beyond their level", "No", 3, ""),
    ("Troubleshooting and Mechanical Skill", "C18", "Identifies likely root cause of common stops", "Can explain the likely cause of a common stop and take the correct first response", "No", 3, ""),
    ("Changeover and Setup", "C19", "Performs routine format change steps", "Completes routine changeover tasks in the correct sequence with minimal correction", "Yes", 3, ""),
    ("Changeover and Setup", "C20", "Verifies correct parts and settings after changeover", "Confirms parts and settings are correct before startup after a changeover", "Yes", 3, ""),
    ("Documentation and Systems", "C21", "Completes work orders accurately", "Enters required work order information accurately and completely", "No", 3, ""),
    ("Documentation and Systems", "C22", "Uses Informance correctly including downtime tracking", "Uses correct downtime codes and enters data accurately without rework", "No", 3, ""),
    ("Documentation and Systems", "C23", "Updates whiteboards and batch information accurately", "Keeps shift and production information accurate and current", "No", 3, ""),
    ("Sanitation and Shutdown", "C24", "Performs proper shutdown", "Shuts down the line in the correct sequence and leaves equipment in the proper state", "No", 3, ""),
    ("Sanitation and Shutdown", "C25", "Leaves the line clean and ready for next shift or sanitation", "Leaves the area and equipment in acceptable condition for the next step in the process", "No", 3, ""),
    ("Escalation and Ownership", "C26", "Knows when and how to escalate", "Escalates at the right time with clear, useful information", "No", 3, ""),
    ("Escalation and Ownership", "C27", "Gives clear shift handoff and ownership communication", "Provides a complete and accurate handoff without leaving important issues unclear", "No", 3, ""),
]


def _format_roster_name(name: str) -> str:
    if "," in name:
        last, first = [p.strip() for p in name.split(",", 1)]
        return f"{first} {last}".strip()
    return str(name).strip()


def _extract_cell_from_department(department: str) -> str:
    if not department:
        return "Unassigned"
    match = re.search(r"Packaging[-\s]*Cell\s*([0-9]+)", str(department), re.IGNORECASE)
    if match:
        return f"Cell {match.group(1)}"
    return "Unassigned"


def load_line_names(repo_root: Path, max_lines: int = 21) -> Tuple[List[str], List[str], str]:
    source_file = repo_root / LINE_SOURCE_FILE
    wb = load_workbook(source_file, data_only=True)
    ws = wb[wb.sheetnames[0]]

    lines: List[str] = []
    cells_by_line: List[str] = []
    active_cell = ""
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[0]:
            active_cell = str(row[0]).strip()
        if row[1]:
            line_name = str(row[1]).strip()
            lines.append(line_name)
            cells_by_line.append(active_cell or "Cell ?")

    if len(lines) < max_lines:
        lines += [f"Line_{i:02d}" for i in range(len(lines) + 1, max_lines + 1)]
        cells_by_line += ["Cell ?"] * (max_lines - len(cells_by_line))

    return lines[:max_lines], cells_by_line[:max_lines], f"Loaded line names from {LINE_SOURCE_FILE} (Sheet1)."


def load_filtered_employees(repo_root: Path) -> Tuple[List[tuple], str]:
    source_file = repo_root / EMPLOYEE_SOURCE_FILE
    wb = load_workbook(source_file, data_only=True)
    ws = wb[wb.sheetnames[0]]

    filtered = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, title, department, hire_date, *_ = row
        if not name or not title:
            continue
        title_l = str(title).lower()
        dept_l = str(department or "").lower()
        is_packaging = "packaging" in dept_l
        is_plt = "tech production line ii" in title_l
        is_mt2 = "tech machine" in title_l or "operator machine ii" in title_l
        if not is_packaging or not (is_plt or is_mt2):
            continue

        role_family = "PLT_MT2"
        employee_name = _format_roster_name(str(name))
        shift = "TBD"
        cell = _extract_cell_from_department(str(department or ""))
        team_lead = "TBD"
        trainer = "TBD"
        if isinstance(hire_date, datetime):
            hire_date = hire_date.date().isoformat()
        elif hasattr(hire_date, "isoformat"):
            hire_date = hire_date.isoformat()
        else:
            hire_date = ""
        active_status = "Active"
        filtered.append((employee_name, role_family, shift, cell, team_lead, trainer, hire_date, active_status))

    filtered = sorted(filtered, key=lambda x: x[0])
    with_ids = []
    for idx, employee in enumerate(filtered, start=1):
        with_ids.append((f"E{3000 + idx}",) + employee)

    note = (
        f"Loaded employees from {EMPLOYEE_SOURCE_FILE} and filtered to Packaging department with titles "
        "matching PLT ('Tech Production Line II') and MT II proxies ('Tech Machine' + 'Operator Machine II')."
    )
    return with_ids, note


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def add_table(ws, name: str, start_cell: str, end_cell: str):
    table = Table(displayName=name, ref=f"{start_cell}:{end_cell}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def build_workbook():
    repo_root = Path(__file__).resolve().parent
    line_names, line_cells, line_name_note = load_line_names(repo_root)
    employees, employee_note = load_filtered_employees(repo_root)

    wb = Workbook()
    ws_instructions = wb.active
    ws_instructions.title = "Instructions"
    ws_employee = wb.create_sheet("Employee_List")
    ws_skills = wb.create_sheet("Skill_Definitions")
    ws_assess = wb.create_sheet("Core_Skill_Assessments")
    ws_lines = wb.create_sheet("Line_Qualifications")
    ws_dash = wb.create_sheet("Dashboard")
    ws_lists = wb.create_sheet("Lists")

    # Instructions
    ws_instructions["A1"] = "Packaging Skills Matrix & Line Qualification Workbook"
    ws_instructions["A1"].font = Font(size=14, bold=True)
    instruction_text = [
        ("Workbook purpose", "Track packaging core skills, line qualifications, and staffing risk for PLT/MT II."),
        ("Tab guide", "Employee_List = roster; Skill_Definitions = competency standards; Core_Skill_Assessments = person vs core skills; Line_Qualifications = line depth; Dashboard = leadership snapshot; Lists = dropdown/admin data."),
        ("Core skill levels", "1 Awareness, 2 Assisted, 3 Independent (target), 4 Trainer."),
        ("Line qualification levels", "1 Not trained, 2 Can run with help, 3 Can run independently, 4 Can train others."),
        ("Update employees", "HR or area leadership updates Employee_List monthly using roster export. Keep Employee_ID unique. Active_Status controls who appears in assessments."),
        ("Update line names", "Packaging leadership updates Lists > line_name whenever line naming changes. Re-run script to refresh dashboard labels."),
        ("Update skill definitions", "Edit rows in Skill_Definitions table. Required level defaults to 3; set different value only where needed."),
        ("Critical gap logic", "Critical_Gap = Yes when Critical_Flag = Yes and Current_Level is below Required_Level_For_Role."),
        ("Line readiness logic", "Core_Critical_Ready = Yes only when employee has zero critical gaps in Core_Skill_Assessments."),
        ("Dashboard flags", "Red/amber flags indicate low line coverage or critical risk. Low coverage threshold is editable in Lists tab."),
        ("Update ownership note", "Supervisors update Current_Level and line qualifications weekly; team leads maintain comments; department leaders review Dashboard in staff meeting."),
    ]
    r = 3
    for title, body in instruction_text:
        ws_instructions[f"A{r}"] = title
        ws_instructions[f"A{r}"].font = Font(bold=True)
        ws_instructions[f"B{r}"] = body
        ws_instructions[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    ws_instructions.column_dimensions["A"].width = 28
    ws_instructions.column_dimensions["B"].width = 120
    ws_instructions.freeze_panes = "A3"

    # Lists
    ws_lists.append(["List_Name", "Value", "Label"])
    rows = [
        ("core_skill_level", 1, "Awareness"),
        ("core_skill_level", 2, "Assisted"),
        ("core_skill_level", 3, "Independent"),
        ("core_skill_level", 4, "Trainer"),
        ("line_level", 1, "Not trained"),
        ("line_level", 2, "Can run with help"),
        ("line_level", 3, "Can run independently"),
        ("line_level", 4, "Can train others"),
        ("role_family", "PLT_MT2", "PLT + MT II"),
        ("shift", "A", "Shift A"),
        ("shift", "B", "Shift B"),
        ("shift", "C", "Shift C"),
        ("shift", "D", "Shift D"),
        ("cell", "Cell 1", "Cell 1"),
        ("cell", "Cell 2", "Cell 2"),
        ("cell", "Cell 3", "Cell 3"),
        ("active_status", "Active", "Active"),
        ("active_status", "Inactive", "Inactive"),
        ("admin", "Low_Coverage_Threshold", "3"),
    ]
    for line, cell in zip(line_names, line_cells):
        rows.append(("line_name", line, f"{line} ({cell})"))
    for row in rows:
        ws_lists.append(row)
    style_header(ws_lists)
    add_table(ws_lists, "ListsTable", "A1", f"C{ws_lists.max_row}")
    ws_lists.column_dimensions["A"].width = 22
    ws_lists.column_dimensions["B"].width = 28
    ws_lists.column_dimensions["C"].width = 36

    # Named ranges for validation lists
    wb.defined_names.add(DefinedName(name="RoleFamilyList", attr_text="Lists!$B$10:$B$10"))
    wb.defined_names.add(DefinedName(name="ShiftList", attr_text="Lists!$B$11:$B$14"))
    wb.defined_names.add(DefinedName(name="CellList", attr_text="Lists!$B$15:$B$17"))
    wb.defined_names.add(DefinedName(name="ActiveStatusList", attr_text="Lists!$B$18:$B$19"))
    wb.defined_names.add(DefinedName(name="CoreLevelList", attr_text="Lists!$B$2:$B$5"))
    wb.defined_names.add(DefinedName(name="LineLevelList", attr_text="Lists!$B$6:$B$9"))
    wb.defined_names.add(DefinedName(name="LowCoverageThreshold", attr_text="Lists!$C$20"))
    wb.defined_names.add(DefinedName(name="LineNamesList", attr_text=f"Lists!$B$21:$B${20 + len(line_names)}"))

    # Employee list
    employee_headers = ["Employee_ID", "Employee_Name", "Role_Family", "Shift", "Cell", "Team_Lead", "Primary_Trainer", "Hire_Date", "Active_Status"]
    ws_employee.append(employee_headers)
    for e in employees:
        ws_employee.append(list(e))
    style_header(ws_employee)
    add_table(ws_employee, "EmployeeTable", "A1", f"I{ws_employee.max_row}")
    ws_employee.freeze_panes = "A2"
    for c, w in {"A": 14, "B": 22, "C": 14, "D": 8, "E": 10, "F": 16, "G": 16, "H": 12, "I": 12}.items():
        ws_employee.column_dimensions[c].width = w
    for col, rng in [("C", "RoleFamilyList"), ("D", "ShiftList"), ("E", "CellList"), ("I", "ActiveStatusList")]:
        dv = DataValidation(type="list", formula1=f"={rng}", allow_blank=False)
        ws_employee.add_data_validation(dv)
        dv.add(f"{col}2:{col}500")

    # Skill definitions
    skills_headers = ["Category", "Competency_ID", "Competency", "Level_3_Standard", "Critical_Flag", "Required_Level_For_Role", "Notes"]
    ws_skills.append(skills_headers)
    for skill in SKILLS:
        ws_skills.append(list(skill))
    style_header(ws_skills)
    add_table(ws_skills, "SkillTable", "A1", f"G{ws_skills.max_row}")
    ws_skills.freeze_panes = "A2"
    for col, w in {"A": 30, "B": 14, "C": 50, "D": 70, "E": 12, "F": 22, "G": 24}.items():
        ws_skills.column_dimensions[col].width = w
    for row in ws_skills.iter_rows(min_row=2, max_row=ws_skills.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Core skill assessments
    assess_headers = [
        "Employee_ID", "Employee_Name", "Role_Family", "Shift", "Cell", "Competency_ID", "Category", "Competency",
        "Level_3_Standard", "Critical_Flag", "Required_Level_For_Role", "Current_Level", "Meets_Required", "Critical_Gap",
        "Last_Assessed_By", "Last_Assessed_Date", "Comments"
    ]
    ws_assess.append(assess_headers)
    active_employees = [e for e in employees if e[8] == "Active"]
    for emp in active_employees:
        for skill in SKILLS:
            ws_assess.append([
                emp[0], emp[1], emp[2], emp[3], emp[4], skill[1], skill[0], skill[2], skill[3], skill[4], skill[5], 2,
                "", "", "", "", ""
            ])
    for r in range(2, ws_assess.max_row + 1):
        ws_assess[f"M{r}"] = f'=IF(L{r}>=K{r},"Yes","No")'
        ws_assess[f"N{r}"] = f'=IF(AND(J{r}="Yes",L{r}<K{r}),"Yes","No")'
    style_header(ws_assess)
    add_table(ws_assess, "AssessmentTable", "A1", f"Q{ws_assess.max_row}")
    ws_assess.freeze_panes = "A2"
    widths = [12, 20, 12, 8, 8, 14, 28, 42, 60, 10, 16, 12, 12, 12, 16, 14, 24]
    for i, w in enumerate(widths, 1):
        ws_assess.column_dimensions[get_column_letter(i)].width = w
    for row in ws_assess.iter_rows(min_row=2, max_row=ws_assess.max_row, min_col=7, max_col=17):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    dv_core = DataValidation(type="list", formula1="=CoreLevelList", allow_blank=False)
    ws_assess.add_data_validation(dv_core)
    dv_core.add(f"L2:L{ws_assess.max_row}")
    # CF
    ws_assess.conditional_formatting.add(
        f"L2:L{ws_assess.max_row}",
        FormulaRule(formula=["$L2<$K2"], fill=PatternFill("solid", fgColor="FCE4D6"))
    )
    ws_assess.conditional_formatting.add(
        f"A2:Q{ws_assess.max_row}",
        FormulaRule(formula=["$N2=\"Yes\""], fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006", bold=True))
    )

    # Line qualifications
    line_headers = ["Employee_ID", "Employee_Name", "Role_Family", "Shift", "Cell", "Core_Critical_Ready"] + line_names + ["Total_Lines_Level_3_Plus", "Total_Lines_Level_4", "Notes"]
    ws_lines.append(line_headers)
    for emp in active_employees:
        ws_lines.append([emp[0], emp[1], emp[2], emp[3], emp[4], ""] + [1] * 21 + [0, 0, ""])
    for r in range(2, ws_lines.max_row + 1):
        ws_lines[f"F{r}"] = f'=IF(COUNTIFS(Core_Skill_Assessments!$A:$A,A{r},Core_Skill_Assessments!$N:$N,"Yes")=0,"Yes","No")'
        ws_lines[f"AB{r}"] = f"=COUNTIF(G{r}:AA{r},\">=3\")"
        ws_lines[f"AC{r}"] = f"=COUNTIF(G{r}:AA{r},4)"
    style_header(ws_lines)
    add_table(ws_lines, "LineQualTable", "A1", f"AD{ws_lines.max_row}")
    ws_lines.freeze_panes = "A2"
    ws_lines.page_setup.orientation = "landscape"
    ws_lines.page_setup.fitToWidth = 1
    ws_lines.page_setup.fitToHeight = 0
    for i in range(1, 31):
        ws_lines.column_dimensions[get_column_letter(i)].width = 11 if 7 <= i <= 27 else 14
    ws_lines.column_dimensions["B"].width = 20
    dv_line = DataValidation(type="list", formula1="=LineLevelList", allow_blank=False)
    ws_lines.add_data_validation(dv_line)
    dv_line.add(f"G2:AA{ws_lines.max_row}")
    ws_lines.conditional_formatting.add(f"G2:AA{ws_lines.max_row}", CellIsRule(operator="equal", formula=["1"], fill=PatternFill("solid", fgColor="F8CBAD")))
    ws_lines.conditional_formatting.add(f"G2:AA{ws_lines.max_row}", CellIsRule(operator="equal", formula=["2"], fill=PatternFill("solid", fgColor="FCE4D6")))
    ws_lines.conditional_formatting.add(f"G2:AA{ws_lines.max_row}", CellIsRule(operator="equal", formula=["3"], fill=PatternFill("solid", fgColor="E2F0D9")))
    ws_lines.conditional_formatting.add(f"G2:AA{ws_lines.max_row}", CellIsRule(operator="equal", formula=["4"], fill=PatternFill("solid", fgColor="C6E0B4")))
    ws_lines.conditional_formatting.add(
        f"G2:AA{ws_lines.max_row}",
        FormulaRule(formula=["=AND($F2=\"No\",G2>=3)"], fill=PatternFill("solid", fgColor="FFD966"), font=Font(bold=True, color="7F6000"))
    )

    # Dashboard
    ws_dash["A1"] = "Packaging Skills Matrix Dashboard"
    ws_dash["A1"].font = Font(size=14, bold=True)
    ws_dash["A2"] = f"Generated: {date.today().isoformat()}"

    ws_dash["A4"] = "Section A: Line Coverage Summary"
    ws_dash["A4"].font = Font(bold=True)
    ws_dash.append(["Line", "Employees Level 3+", "Employees Level 4", "Low Coverage Flag"])
    header_row = ws_dash.max_row
    for idx in range(1, 22):
        row = ws_dash.max_row + 1
        line_col = get_column_letter(6 + idx)
        ws_dash[f"A{row}"] = f"=Lists!B{20 + idx}"
        ws_dash[f"B{row}"] = f"=COUNTIF(Line_Qualifications!{line_col}:{line_col},\">=3\")"
        ws_dash[f"C{row}"] = f"=COUNTIF(Line_Qualifications!{line_col}:{line_col},4)"
        ws_dash[f"D{row}"] = f"=IF(B{row}<LowCoverageThreshold,\"LOW\",\"OK\")"

    # Compact KPI strip
    ws_dash["F4"] = "KPI"
    ws_dash["G4"] = "Value"
    ws_dash["F5"] = "Active PLT/MT II"
    ws_dash["G5"] = "=COUNTA(Employee_List!A:A)-1"
    ws_dash["F6"] = "Critical-ready"
    ws_dash["G6"] = '=COUNTIF(Line_Qualifications!F:F,"Yes")'
    ws_dash["F7"] = "Lines flagged LOW"
    ws_dash["G7"] = f'=COUNTIF(D{header_row+1}:D{header_row+21},"LOW")'

    # Section B
    start_b = ws_dash.max_row + 2
    ws_dash[f"A{start_b}"] = "Section B: Critical Skill Gaps by Employee"
    ws_dash[f"A{start_b}"].font = Font(bold=True)
    ws_dash[f"A{start_b+1}"] = "Employee_Name"
    ws_dash[f"B{start_b+1}"] = "Critical_Gaps"
    ws_dash[f"C{start_b+1}"] = "All_Skills_Below_Required"
    for i, emp in enumerate(active_employees, start=start_b + 2):
        ws_dash[f"A{i}"] = emp[1]
        ws_dash[f"B{i}"] = f'=COUNTIFS(Core_Skill_Assessments!$B:$B,A{i},Core_Skill_Assessments!$N:$N,"Yes")'
        ws_dash[f"C{i}"] = f'=COUNTIFS(Core_Skill_Assessments!$B:$B,A{i},Core_Skill_Assessments!$M:$M,"No")'

    # Section C
    start_c = ws_dash.max_row + 2
    ws_dash[f"A{start_c}"] = "Section C: Critical Skill Gaps by Competency"
    ws_dash[f"A{start_c}"].font = Font(bold=True)
    ws_dash[f"A{start_c+1}"] = "Competency_ID"
    ws_dash[f"B{start_c+1}"] = "Competency"
    ws_dash[f"C{start_c+1}"] = "Employees_Below_Required"
    crit_skills = [s for s in SKILLS if s[4] == "Yes"]
    for i, skill in enumerate(crit_skills, start=start_c + 2):
        ws_dash[f"A{i}"] = skill[1]
        ws_dash[f"B{i}"] = skill[2]
        ws_dash[f"C{i}"] = f'=COUNTIFS(Core_Skill_Assessments!$F:$F,A{i},Core_Skill_Assessments!$N:$N,"Yes")'

    # Section D
    start_d = ws_dash.max_row + 2
    ws_dash[f"A{start_d}"] = "Section D: Cross Training Opportunity View"
    ws_dash[f"A{start_d}"].font = Font(bold=True)
    ws_dash[f"A{start_d+1}"] = "Employee_Name"
    ws_dash[f"B{start_d+1}"] = "Core_Critical_Ready"
    ws_dash[f"C{start_d+1}"] = "Lines_Level_3_Plus"
    row_d = start_d + 2
    for idx in range(2, ws_lines.max_row + 1):
        ws_dash[f"A{row_d}"] = f"=Line_Qualifications!B{idx}"
        ws_dash[f"B{row_d}"] = f"=Line_Qualifications!F{idx}"
        ws_dash[f"C{row_d}"] = f"=Line_Qualifications!AB{idx}"
        row_d += 1

    # Section E
    start_e = ws_dash.max_row + 2
    ws_dash[f"A{start_e}"] = "Section E: Trainer Bench Strength"
    ws_dash[f"A{start_e}"].font = Font(bold=True)
    ws_dash[f"A{start_e+1}"] = "Employee_Name"
    ws_dash[f"B{start_e+1}"] = "Level_4_Line_Qualifications"
    for i, idx in enumerate(range(2, ws_lines.max_row + 1), start=start_e + 2):
        ws_dash[f"A{i}"] = f"=Line_Qualifications!B{idx}"
        ws_dash[f"B{i}"] = f"=Line_Qualifications!AC{idx}"

    style_header(ws_dash, row=header_row)
    style_header(ws_dash, row=start_b + 1)
    style_header(ws_dash, row=start_c + 1)
    style_header(ws_dash, row=start_d + 1)
    style_header(ws_dash, row=start_e + 1)
    ws_dash.freeze_panes = "A6"
    ws_dash.column_dimensions["A"].width = 30
    ws_dash.column_dimensions["B"].width = 18
    ws_dash.column_dimensions["C"].width = 18
    ws_dash.column_dimensions["D"].width = 16
    ws_dash.column_dimensions["F"].width = 18
    ws_dash.column_dimensions["G"].width = 12
    ws_dash.conditional_formatting.add(f"D{header_row+1}:D{header_row+21}", FormulaRule(formula=[f'$D{header_row+1}="LOW"'], fill=PatternFill("solid", fgColor="FFC7CE")))
    style_header(ws_dash, row=4)

    # common alignment
    for ws in [ws_employee, ws_skills, ws_assess, ws_lines, ws_dash]:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for c in row:
                if c.alignment is None:
                    c.alignment = Alignment(vertical="top")

    # highlight critical fields
    for r in range(2, ws_skills.max_row + 1):
        if ws_skills[f"E{r}"].value == "Yes":
            ws_skills[f"E{r}"].fill = PatternFill("solid", fgColor="FFC7CE")
            ws_skills[f"E{r}"].font = Font(color="9C0006", bold=True)
    ws_assess.conditional_formatting.add(
        f"J2:J{ws_assess.max_row}",
        FormulaRule(formula=['$J2="Yes"'], fill=PatternFill("solid", fgColor="FFF2CC"), font=Font(bold=True))
    )

    ws_lists.sheet_state = "hidden"

    wb.save(repo_root / OUTPUT_FILE)

    # Validation reload
    loaded = load_workbook(repo_root / OUTPUT_FILE)
    tab_names = loaded.sheetnames

    print("Completion summary")
    print(f"- Files created: build_workbook.py, {OUTPUT_FILE}")
    print(f"- Tabs created: {', '.join(tab_names)}")
    print(f"- Assumptions/fallbacks: Role_Family unified as PLT_MT2; shift/team lead/trainer defaulted to TBD; {employee_note}")
    print(f"- Line names source: {line_name_note}")


if __name__ == "__main__":
    build_workbook()
